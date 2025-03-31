# Copyright 2023 agenius666
# GitHub: https://github.com/agenius666/Counterparty-Account-Processor
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import win32com.client as win32
import threading


class AdvancedAccountingProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("对方科目 - 1.0.0")

        # 界面组件初始化
        self.create_widgets()
        self.processing = False

    def create_widgets(self):
        # 输入框架
        input_frame = ttk.LabelFrame(self.root, text="处理设置")
        input_frame.pack(padx=10, pady=5, fill="x")

        # 模式选择
        self.mode_var = tk.StringVar(value="separate")  # 默认选择“借贷方分开”
        mode_frame = ttk.Frame(input_frame)
        mode_frame.pack(fill="x", pady=5)
        ttk.Radiobutton(mode_frame, text="借贷方分开", variable=self.mode_var, value="separate", command=self.toggle_mode).pack(side='left')
        ttk.Radiobutton(mode_frame, text="借贷方在一起", variable=self.mode_var, value="together", command=self.toggle_mode).pack(side='left')

        # 借贷方分开模式下的输入
        self.separate_frame = ttk.Frame(input_frame)
        self.separate_frame.pack(fill="x", pady=5)
        fields_separate = [
            ("凭证字号列", "voucher_col"),
            ("科目名称列", "subject_col"),
            ("借方金额列", "debit_col"),
            ("贷方金额列", "credit_col"),
            ("工作表名称", "sheet_name"),
            ("目标列位置", "target_col")
        ]
        self.entries_separate = {}
        for text, var_name in fields_separate:
            row = ttk.Frame(self.separate_frame)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text=text, width=12).pack(side='left')
            entry = ttk.Entry(row)
            entry.pack(side='right', expand=True, fill="x")
            self.entries_separate[var_name] = entry

        # 借贷方在一起模式下的输入
        self.together_frame = ttk.Frame(input_frame)
        fields_together = [
            ("凭证字号列", "voucher_col"),
            ("科目名称列", "subject_col"),
            ("借贷方金额列", "amount_col"),
            ("借贷标识列", "direction_col"),
            ("借方标识", "debit_flag"),
            ("贷方标识", "credit_flag"),
            ("贷方处理方式", "credit_action"),
            ("工作表名称", "sheet_name"),
            ("目标列位置", "target_col")
        ]
        self.entries_together = {}
        for text, var_name in fields_together:
            row = ttk.Frame(self.together_frame)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text=text, width=12).pack(side='left')
            if var_name == "credit_action":
                combobox = ttk.Combobox(row, values=["直接等于", "取相反数"], state="readonly")
                combobox.pack(side='right', expand=True, fill="x")
                self.entries_together[var_name] = combobox
            else:
                entry = ttk.Entry(row)
                entry.pack(side='right', expand=True, fill="x")
                self.entries_together[var_name] = entry

        # 默认显示“借贷方分开”界面
        self.toggle_mode()

        # 文件操作模式
        file_mode_frame = ttk.Frame(self.root)
        file_mode_frame.pack(padx=10, pady=5, fill="x")
        self.file_mode_var = tk.StringVar(value="single")
        ttk.Radiobutton(file_mode_frame, text="单文件处理", variable=self.file_mode_var, value="single").pack(side='left')
        ttk.Radiobutton(file_mode_frame, text="批量处理", variable=self.file_mode_var, value="batch").pack(side='left')

        # 文件选择
        self.file_path = tk.StringVar()
        self.dir_path = tk.StringVar()
        file_frame = ttk.Frame(self.root)
        file_frame.pack(fill="x", padx=10, pady=5)
        ttk.Button(file_frame, text="选择文件/目录", command=self.select_path).pack(side='left')
        ttk.Label(file_frame, textvariable=self.file_path).pack(side='left', padx=5)

        # 保存路径
        self.save_path = tk.StringVar()
        save_frame = ttk.Frame(self.root)
        save_frame.pack(fill="x", padx=10, pady=5)
        ttk.Button(save_frame, text="选择保存目录", command=self.select_save_path).pack(side='left')
        ttk.Label(save_frame, textvariable=self.save_path).pack(side='left', padx=5)

        # 进度条
        self.progress = ttk.Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(pady=10)

        # 控制按钮
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="开始处理", command=self.start_processing).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="停止", command=self.stop_processing).pack(side='right')

    def toggle_mode(self):
        """切换模式界面"""
        if self.mode_var.get() == "separate":
            self.together_frame.pack_forget()
            self.separate_frame.pack(fill="x", pady=5)
        else:
            self.separate_frame.pack_forget()
            self.together_frame.pack(fill="x", pady=5)

    def select_path(self):
        if self.file_mode_var.get() == "single":
            filepath = filedialog.askopenfilename(
                filetypes=[("Excel文件", "*.xls*"), ("All files", "*.*")])
            if filepath:
                self.file_path.set(filepath)
        else:
            dirpath = filedialog.askdirectory()
            if dirpath:
                self.dir_path.set(dirpath)
                self.file_path.set(f"已选择目录：{dirpath}")

    def select_save_path(self):
        path = filedialog.askdirectory()
        if path:
            self.save_path.set(path)

    def start_processing(self):
        if self.processing:
            return

        # 检查必填字段
        if self.mode_var.get() == "separate":
            required_fields = ["voucher_col", "subject_col", "debit_col", "credit_col"]
            entries = self.entries_separate
        else:
            required_fields = ["voucher_col", "subject_col", "amount_col", "direction_col", "debit_flag", "credit_flag"]
            entries = self.entries_together

        if not all(entries[e].get() for e in required_fields):
            messagebox.showerror("错误", "必填字段不能为空")
            return

        self.processing = True
        thread = threading.Thread(target=self.process_files)
        thread.start()

    def stop_processing(self):
        self.processing = False

    def process_files(self):
        try:
            # 获取处理参数
            if self.mode_var.get() == "separate":
                params = {
                    "voucher_col": self.entries_separate["voucher_col"].get().strip(),
                    "subject_col": self.entries_separate["subject_col"].get().strip(),
                    "debit_col": self.entries_separate["debit_col"].get().strip(),
                    "credit_col": self.entries_separate["credit_col"].get().strip(),
                    "sheet_name": self.entries_separate["sheet_name"].get().strip() or None,
                    "target_col": self.entries_separate["target_col"].get().strip().upper()
                }
            else:
                params = {
                    "voucher_col": self.entries_together["voucher_col"].get().strip(),
                    "subject_col": self.entries_together["subject_col"].get().strip(),
                    "amount_col": self.entries_together["amount_col"].get().strip(),
                    "direction_col": self.entries_together["direction_col"].get().strip(),
                    "debit_flag": self.entries_together["debit_flag"].get().strip(),
                    "credit_flag": self.entries_together["credit_flag"].get().strip(),
                    "credit_action": self.entries_together["credit_action"].get(),
                    "sheet_name": self.entries_together["sheet_name"].get().strip() or None,
                    "target_col": self.entries_together["target_col"].get().strip().upper()
                }

            # 获取文件列表
            if self.file_mode_var.get() == "single":
                files = [self.file_path.get()]
            else:
                dir_path = self.dir_path.get()
                files = [os.path.join(dir_path, f) for f in os.listdir(dir_path)
                         if f.lower().endswith(('.xls', '.xlsx', '.xlsm'))]

            total_files = len(files)
            for i, file_path in enumerate(files):
                if not self.processing:
                    break

                try:
                    if self.mode_var.get() == "separate":
                        self.process_separate_mode(file_path, params, i + 1, total_files)
                    else:
                        self.process_together_mode(file_path, params, i + 1, total_files)
                except Exception as e:
                    messagebox.showerror("错误", f"处理文件失败：{file_path}\n{str(e)}")

            messagebox.showinfo("完成", "处理完成！" if self.processing else "处理已中止")

        finally:
            self.processing = False
            self.root.config(cursor="")
            self.progress["value"] = 0

    def remove_filters_with_win32com(self, file_path, sheet_name=None):
        """使用 win32com 取消 Excel 文件中的筛选"""
        try:
            # 启动 Excel 应用程序
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False  # 不显示 Excel 界面
            excel.DisplayAlerts = False  # 禁用警告

            # 打开工作簿
            wb = excel.Workbooks.Open(file_path)
            if sheet_name:
                ws = wb.Sheets(sheet_name)
            else:
                ws = wb.Sheets(1)  # 默认选择第一个工作表

            # 取消筛选
            if ws.AutoFilterMode:
                ws.AutoFilterMode = False

            # 保存并关闭工作簿
            wb.Save()
            wb.Close()
            excel.Quit()

        except Exception as e:
            print(f"使用 win32com 取消筛选失败：{str(e)}")
            raise e
    def process_separate_mode(self, file_path, params, current_num, total_files):
        """处理借贷方分开模式"""
        try:
            # 使用 win32com 取消筛选
            self.remove_filters_with_win32com(file_path, params["sheet_name"])

            # 更新进度
            self.root.config(cursor="wait")
            progress = int((current_num / total_files) * 100)
            self.progress["value"] = progress
            self.root.update_idletasks()

            # 读取文件
            if file_path.endswith('.xls'):
                df = pd.read_excel(file_path, sheet_name=params["sheet_name"], engine='xlrd')
            else:
                wb = load_workbook(file_path, read_only=False, keep_vba=True, keep_links=False)
                sheet_name = params["sheet_name"] or wb.sheetnames[0]
                ws = wb[sheet_name]

                data = ws.values
                columns = next(data)
                df = pd.DataFrame(data, columns=columns)
                wb.close()

            def excel_column_to_num(col_str):
                """将Excel列字母转换为数字索引（如'A'->1, 'AI'->35）"""
                num = 0
                for i, c in enumerate(reversed(col_str.upper())):
                    num += (ord(c) - 64) * (26 ** i)
                return num

            # 获取列名
            def get_column_name(df, col_input):
                if col_input.isalpha():
                    col_idx = excel_column_to_num(col_input)  # 使用新函数
                    return df.columns[col_idx - 1]
                elif col_input.isdigit():
                    col_idx = int(col_input)
                    return df.columns[col_idx - 1]
                else:
                    return col_input

            voucher_col = get_column_name(df, params["voucher_col"])
            debit_col = get_column_name(df, params["debit_col"])
            credit_col = get_column_name(df, params["credit_col"])
            subject_col = get_column_name(df, params["subject_col"])

            # 检查列名是否存在
            required_columns = [voucher_col, debit_col, credit_col, subject_col]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messagebox.showerror("错误", f"以下列不存在于文件中：{', '.join(missing_columns)}")
                return

            # 填充借方列和贷方列中的空值为 0
            df[debit_col] = df[debit_col].fillna(0).astype(float)
            df[credit_col] = df[credit_col].fillna(0).astype(float)

            # 确保凭证字号列为字符串格式，并去掉 .0 后缀
            df[voucher_col] = df[voucher_col].astype(str).apply(lambda x: x[:-2] if x.endswith('.0') else x)

            # 清理 voucher_col 列
            df[voucher_col] = df[voucher_col].astype(str).str.strip()

            # 初始化结果列表
            result_list = [""] * len(df)  # 创建一个与原始数据行数相同的空列表

            # 按凭证字号分组
            grouped = df.groupby(voucher_col)

            # 遍历分组，计算每个凭证字号对应的对方科目
            for voucher, group in grouped:
                debit_subjects = []
                credit_subjects = []

                # 收集借方科目和贷方科目
                for index, row in group.iterrows():
                    subject = row[subject_col]
                    debit = row[debit_col]
                    credit = row[credit_col]
                    if debit > 0:
                        debit_subjects.append(subject)
                    elif credit > 0:
                        credit_subjects.append(subject)

                # 去重并转换为字符串
                debit_subjects = list(set(debit_subjects))
                credit_subjects = list(set(credit_subjects))
                debit_subjects_str = "、".join(debit_subjects)
                credit_subjects_str = "、".join(credit_subjects)

                # 根据每行的借贷方向，决定写入的内容
                for index, row in group.iterrows():
                    if row[debit_col] > 0:
                        result_list[df.index.get_loc(index)] = credit_subjects_str
                    elif row[credit_col] > 0:
                        result_list[df.index.get_loc(index)] = debit_subjects_str

            # 保存文件
            save_dir = self.save_path.get() or os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            name, ext = os.path.splitext(base_name)
            save_path = os.path.join(save_dir, f"{name}_处理后{ext}")

            if file_path.endswith('.xls'):
                # 保存为 .xlsx
                save_path = save_path.replace('.xls', '.xlsx')
                self.save_normal_file(file_path, save_path, result_list, params)
            else:
                # 使用 openpyxl 保存，保留格式和 VBA
                self.save_with_vba(file_path, save_path, result_list, params)

        except Exception as e:
            print(f"处理文件失败：{file_path}")  # 调试信息
            print(f"错误详情：{str(e)}")  # 调试信息
            messagebox.showerror("错误", f"处理文件失败：{file_path}\n{str(e)}")

    def process_together_mode(self, file_path, params, current_num, total_files):
        """处理借贷方在一起模式"""
        try:
            # 使用 win32com 取消筛选
            self.remove_filters_with_win32com(file_path, params["sheet_name"])

            # 更新进度
            self.root.config(cursor="wait")
            progress = int((current_num / total_files) * 100)
            self.progress["value"] = progress
            self.root.update_idletasks()

            # 读取文件
            if file_path.endswith('.xls'):
                df = pd.read_excel(file_path, sheet_name=params["sheet_name"], engine='xlrd')
            else:
                wb = load_workbook(file_path, read_only=False, keep_vba=True, keep_links=False)
                sheet_name = params["sheet_name"] or wb.sheetnames[0]
                ws = wb[sheet_name]

                data = ws.values
                columns = next(data)
                df = pd.DataFrame(data, columns=columns)
                wb.close()

            def excel_column_to_num(col_str):
                """将Excel列字母转换为数字索引（如'A'->1, 'AI'->35）"""
                num = 0
                for i, c in enumerate(reversed(col_str.upper())):
                    num += (ord(c) - 64) * (26 ** i)
                return num

            # 获取列名
            def get_column_name(df, col_input):
                if col_input.isalpha():
                    col_idx = excel_column_to_num(col_input)  # 使用新函数
                    return df.columns[col_idx - 1]
                elif col_input.isdigit():
                    col_idx = int(col_input)
                    return df.columns[col_idx - 1]
                else:
                    return col_input

            voucher_col = get_column_name(df, params["voucher_col"])
            subject_col = get_column_name(df, params["subject_col"])
            amount_col = get_column_name(df, params["amount_col"])
            direction_col = get_column_name(df, params["direction_col"])

            # 检查列名是否存在
            required_columns = [voucher_col, subject_col, amount_col, direction_col]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messagebox.showerror("错误", f"以下列不存在于文件中：{', '.join(missing_columns)}")
                return

            # 填充金额列中的空值为 0
            df[amount_col] = df[amount_col].fillna(0).astype(float)

            # 确保凭证字号列为字符串格式，并去掉 .0 后缀
            df[voucher_col] = df[voucher_col].astype(str).apply(lambda x: x[:-2] if x.endswith('.0') else x)

            # 清理 voucher_col 列
            df[voucher_col] = df[voucher_col].astype(str).str.strip()

            # 初始化结果列表
            result_list = [""] * len(df)  # 创建一个与原始数据行数相同的空列表

            # 按凭证字号分组
            grouped = df.groupby(voucher_col)

            # 遍历分组，计算每个凭证字号对应的对方科目
            for voucher, group in grouped:
                debit_subjects = []
                credit_subjects = []

                # 收集借方科目和贷方科目
                for index, row in group.iterrows():
                    subject = row[subject_col]
                    amount = row[amount_col]
                    direction = row[direction_col]
                    if direction == params["debit_flag"]:
                        debit_subjects.append(subject)
                    elif direction == params["credit_flag"]:
                        credit_subjects.append(subject)

                # 去重并转换为字符串
                debit_subjects = list(set(debit_subjects))
                credit_subjects = list(set(credit_subjects))
                debit_subjects_str = "、".join(debit_subjects)
                credit_subjects_str = "、".join(credit_subjects)

                # 根据每行的借贷方向，决定写入的内容
                for index, row in group.iterrows():
                    direction = row[direction_col]
                    if direction == params["debit_flag"]:
                        result_list[df.index.get_loc(index)] = credit_subjects_str
                    elif direction == params["credit_flag"]:
                        result_list[df.index.get_loc(index)] = debit_subjects_str

            # 保存文件
            save_dir = self.save_path.get() or os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            name, ext = os.path.splitext(base_name)
            save_path = os.path.join(save_dir, f"{name}_处理后{ext}")

            if file_path.endswith('.xls'):
                # 保存为 .xlsx
                save_path = save_path.replace('.xls', '.xlsx')
                self.save_normal_file(file_path, save_path, result_list, params)
            else:
                # 使用 openpyxl 保存，保留格式和 VBA
                self.save_with_vba(file_path, save_path, result_list, params)

        except Exception as e:
            print(f"处理文件失败：{file_path}")  # 调试信息
            print(f"错误详情：{str(e)}")  # 调试信息
            messagebox.showerror("错误", f"处理文件失败：{file_path}\n{str(e)}")

    def save_normal_file(self, src_path, dst_path, result_list, params):
        # 使用 openpyxl 处理非宏文件
        wb = load_workbook(src_path)
        sheet_name = params["sheet_name"] or wb.sheetnames[0]
        ws = wb[sheet_name]

        def excel_column_to_num(col_str):
            """将Excel列字母转换为数字索引（如'A'->1, 'AI'->35）"""
            num = 0
            for i, c in enumerate(reversed(col_str.upper())):
                num += (ord(c) - 64) * (26 ** i)
            return num

        # 在目标列写入数据
        target_col = params["target_col"]
        if target_col.isalpha():
            col_idx = excel_column_to_num(target_col)
        else:
            col_idx = int(target_col)

        # 写入标题
        ws.cell(row=1, column=col_idx, value="对方科目")
        ws.cell(1, col_idx).font = Font(bold=True)
        ws.cell(1, col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for idx, value in enumerate(result_list, start=2):
            ws.cell(row=idx, column=col_idx, value=value)

        # 保存文件
        wb.save(dst_path)

    def save_with_vba(self, src_path, dst_path, result_list, params):
        # 使用 openpyxl 处理带宏文件
        wb = load_workbook(src_path, keep_vba=True)
        sheet_name = params["sheet_name"] or wb.sheetnames[0]
        ws = wb[sheet_name]

        def excel_column_to_num(col_str):
            """将Excel列字母转换为数字索引（如'A'->1, 'AI'->35）"""
            num = 0
            for i, c in enumerate(reversed(col_str.upper())):
                num += (ord(c) - 64) * (26 ** i)
            return num

        # 在目标列写入数据
        target_col = params["target_col"]
        if target_col.isalpha():
            col_idx = excel_column_to_num(target_col)
        else:
            col_idx = int(target_col)

        # 写入标题
        ws.cell(row=1, column=col_idx, value="对方科目")
        ws.cell(1, col_idx).font = Font(bold=True)
        ws.cell(1, col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for idx, value in enumerate(result_list, start=2):
            ws.cell(row=idx, column=col_idx, value=value)

        # 保存文件
        wb.save(dst_path)


if __name__ == "__main__":
    root = tk.Tk()
    app = AdvancedAccountingProcessor(root)
    root.mainloop()
